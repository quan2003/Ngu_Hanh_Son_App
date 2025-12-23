#!/usr/bin/env python3
"""Create Excel template for administrative units data entry."""

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    print("Installing required packages...")
    import subprocess

    subprocess.check_call(["pip", "install", "pandas", "openpyxl"])
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

# Create workbook
wb = Workbook()

# Sheet 1: Chi bộ
ws_cb = wb.active
ws_cb.title = "Chi bộ"

# Headers
headers_cb = [
    "STT",
    "Tên chi bộ",
    "Mô tả",
    "Tọa độ trung tâm (lat,lng)",
    "Lãnh đạo",
    "Số điện thoại",
    "Số thành viên",
    "Màu hiển thị",
]

# Add headers with style
for col, header in enumerate(headers_cb, 1):
    cell = ws_cb.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add example data
example_cb = [
    [
        1,
        "Chi bộ Khuê Đông",
        "Chi bộ Khuê Đông - Phường Ngũ Hành Sơn",
        "16.0530,108.2020",
        "Nguyễn Văn A",
        "0123456789",
        150,
        "#2196F3",
    ],
    [
        2,
        "CB 1A",
        "Chi bộ 1A - Khu vực A",
        "16.0540,108.2030",
        "Trần Thị B",
        "0987654321",
        120,
        "#1976D2",
    ],
    [
        3,
        "CB 2A",
        "Chi bộ 2A - Khu vực B",
        "16.0520,108.2040",
        "Lê Văn C",
        "0912345678",
        100,
        "#1565C0",
    ],
]

for row_data in example_cb:
    ws_cb.append(row_data)

# Adjust column widths
ws_cb.column_dimensions["B"].width = 25
ws_cb.column_dimensions["C"].width = 35
ws_cb.column_dimensions["D"].width = 25
ws_cb.column_dimensions["E"].width = 20
ws_cb.column_dimensions["F"].width = 15
ws_cb.column_dimensions["I"].width = 15

# Sheet 2: Tổ dân phố
ws_tdp = wb.create_sheet("Tổ dân phố")

headers_tdp = [
    "STT",
    "Tên tổ",
    "Chi bộ",
    "Mô tả",
    "Tọa độ trung tâm (lat,lng)",
    "Lãnh đạo",
    "Số điện thoại",
    "Số hộ dân",
    "Màu hiển thị",
]

# Add headers
for col, header in enumerate(headers_tdp, 1):
    cell = ws_tdp.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add example data
example_tdp = [
    [
        1,
        "Tổ 230",
        "Chi bộ Khuê Đông",
        "Tổ dân phố 230",
        "16.0535,108.2025",
        "Phạm Văn D",
        "0901234567",
        50,
        "#FF9800",
    ],
    [
        2,
        "Tổ 259",
        "Chi bộ Khuê Đông",
        "Tổ dân phố 259",
        "16.0545,108.2035",
        "Hoàng Thị E",
        "0923456789",
        45,
        "#F57C00",
    ],
    [
        3,
        "Tổ 258",
        "Chi bộ Khuê Đông",
        "Tổ dân phố 258",
        "16.0525,108.2015",
        "Ngô Văn F",
        "0934567890",
        40,
        "#E65100",
    ],
    [
        4,
        "Tổ 66",
        "CB 1A",
        "Tổ dân phố 66",
        "16.0555,108.2045",
        "Đặng Thị G",
        "0945678901",
        35,
        "#EF6C00",
    ],
]

for row_data in example_tdp:
    ws_tdp.append(row_data)

# Adjust column widths
ws_tdp.column_dimensions["B"].width = 20
ws_tdp.column_dimensions["C"].width = 25
ws_tdp.column_dimensions["D"].width = 30
ws_tdp.column_dimensions["E"].width = 25
ws_tdp.column_dimensions["F"].width = 20
ws_tdp.column_dimensions["G"].width = 15
ws_tdp.column_dimensions["J"].width = 15

# Sheet 3: Hướng dẫn
ws_guide = wb.create_sheet("Hướng dẫn")

guide_text = [
    ["HƯỚNG DẪN NHẬP LIỆU"],
    [""],
    ["1. NHẬP THÔNG TIN CHI BỘ (Sheet 'Chi bộ'):"],
    ["   - STT: Số thứ tự tự động"],
    ["   - Tên chi bộ: Tên đầy đủ của chi bộ"],
    ["   - Mô tả: Mô tả chi tiết về chi bộ"],
    ["   - Tọa độ: Định dạng 'latitude,longitude' (VD: 16.0530,108.2020)"],
    ["   - Lãnh đạo: Họ tên lãnh đạo chi bộ"],
    ["   - Số điện thoại: Số điện thoại liên lạc"],
    ["   - Số thành viên: Tổng số đảng viên trong chi bộ"],
    ["   - Màu hiển thị: Mã màu hex (VD: #2196F3 - xanh dương)"],
    [""],
    ["2. NHẬP THÔNG TIN TỔ DÂN PHỐ (Sheet 'Tổ dân phố'):"],
    ["   - STT: Số thứ tự tự động"],
    ["   - Tên tổ: Tên tổ dân phố (VD: Tổ 230)"],
    ["   - Chi bộ: Tên chi bộ quản lý (phải trùng với tên trong sheet Chi bộ)"],
    ["   - Mô tả: Mô tả về khu vực"],
    ["   - Tọa độ: Định dạng 'latitude,longitude'"],
    ["   - Lãnh đạo: Họ tên tổ trưởng"],
    ["   - Số điện thoại: Số điện thoại liên lạc"],
    ["   - Số hộ dân: Số hộ gia đình trong tổ"],
    ["   - Màu hiển thị: Mã màu hex (VD: #FF9800 - cam)"],
    [""],
    ["3. LẤY TỌA ĐỘ:"],
    ["   - Mở Google Maps: https://maps.google.com"],
    ["   - Click chuột phải vào vị trí → Copy coordinates"],
    ["   - Paste vào cột 'Tọa độ trung tâm'"],
    ["   - Hoặc nhìn từ AutoCAD và ước lượng vị trí"],
    [""],
    ["4. MÃ MÀU GỢI Ý:"],
    ["   - Xanh dương: #2196F3, #1976D2, #1565C0"],
    ["   - Cam: #FF9800, #F57C00, #E65100"],
    ["   - Xanh lá: #4CAF50, #388E3C, #2E7D32"],
    ["   - Đỏ: #F44336, #D32F2F, #C62828"],
    ["   - Tím: #9C27B0, #7B1FA2, #6A1B9A"],
    [""],
    ["5. SAU KHI NHẬP XONG:"],
    ["   - Lưu file Excel"],
    ["   - Chạy script: node scripts/upload_excel_to_firebase.js"],
    ["   - Script sẽ tự động import vào Firebase Firestore"],
    [""],
    ["6. LƯU Ý:"],
    ["   - Không xóa dòng tiêu đề (dòng 1)"],
    ["   - Không thay đổi tên các cột"],
    ["   - Tọa độ phải nằm trong phạm vi Đà Nẵng (~16.0°N, 108.2°E)"],
    ["   - Tên chi bộ phải khớp giữa 2 sheet"],
]

for row_idx, row_data in enumerate(guide_text, 1):
    cell = ws_guide.cell(row=row_idx, column=1, value=row_data[0])
    if row_idx == 1:
        cell.font = Font(bold=True, size=14, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="4CAF50", end_color="4CAF50", fill_type="solid"
        )
    elif ":" in str(row_data[0]) and row_data[0][0].isdigit():
        cell.font = Font(bold=True, size=12)

ws_guide.column_dimensions["A"].width = 80

# Save file
output_file = "tools/template_admin_units.xlsx"
wb.save(output_file)

print(f"✅ Excel template created: {output_file}")
print("\nNext steps:")
print("1. Open the Excel file and fill in your data")
print("2. Save the file as: assets/admin_units_data.xlsx")
print("3. Run: python tools/import_excel_to_firebase.py")
